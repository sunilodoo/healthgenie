# -*- coding: utf-8 -*-
import xlwt
import openpyxl
import base64
import calendar
from io import StringIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError, Warning
from datetime import datetime

class XlsReport(models.TransientModel):
    _name = "xls.report"
    _description = 'Xls Report'
    month_of_sale = fields.Selection([
        ('jan', 'January'),
        ('feb', 'February'),
        ('mar', 'March'),
        ('apr', 'April'),
        ('may', 'May'),
        ('jun', 'June'),
        ('jul', 'July'),
        ('aug', 'August'),
        ('sep', 'September'),
        ('oct', 'October'),
        ('nov', 'November'),
        ('dec', 'December')
        ], string="Month of sale", default=False)
    # start_date = fields.Date(string='Date From', required=True)
    # end_date = fields.Date(string='Date To', required=True)
    start_date = fields.Date(string='Date From')
    end_date = fields.Date(string='Date To')
    invoice_state = fields.Selection([
        ('draft', 'Draft'),
        ('posted', 'Posted'),
        ('cancel', 'Cancelled'),
        ], string='Status', default='posted', required=True)
    file_name = fields.Char('Name', size=256)
    file_xls = fields.Binary(' Report', readonly=True)
    flag = fields.Boolean(string="Flag")
    _sql_constraints = [
            ('check','CHECK((start_date <= end_date))',"End date must be greater then start date")  
    ]
    # def action_all_order(self):
    #     inv_ids=self.env['account.move'].search([], order='id asc')
    #     workbook = openpyxl.Workbook()
    #     # wb = workbook.create_sheet('Sheet1')
    #     ws = workbook.active
    #     # s_h = xlwt.easyxf("font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center")
    #     # print("---------------------------------writting into sheet--------------------")
    #     ws.cell(row = 1, column = 1).value = 'Invoice No'
    #     ws.cell(row = 1, column = 2).value = 'Month'
    #     ws.cell(row = 1, column = 3).value = 'Suborder no'
    #     ws.cell(row = 1, column = 4).value = 'Product Id'
    #     ws.cell(row = 1, column = 5).value = 'Product Name'
    #     row = 2
    #     for inv in inv_ids:
    #         for inv_l in inv.invoice_line_ids:
    #             ws.cell(row, 1).value = inv.invoice_number
    #             ws.cell(row, 2).value =  inv.month_of_sale
    #             ws.cell(row, 3).value = inv.order_id
    #             ws.cell(row, 4).value = inv_l.sku_name
    #             ws.cell(row, 5).value = inv_l.product_id.name
    #             row+=1
    #     filename = "/tmp/All_Order.xlsx"
    #     workbook.save(filename)
    #     file = open(filename, "rb")
    #     file_data = file.read()
    #     out = base64.encodebytes(file_data)
    #     self.write({'file_name': filename, 'file_xls':out, 'flag': True})
    #     return {
    #        'type': 'ir.actions.act_window',
    #        'res_model': 'xls.report',
    #        'view_mode': 'form',
    #        'view_type': 'form',
    #        'res_id': self.id,
    #        'target': 'new',
    #     }
    def action_all_order(self):
        inv_ids=self.env['account.move'].search([], order='id asc')
        workbook = openpyxl.Workbook()
        # wb = workbook.create_sheet('Sheet1')
        ws = workbook.active
        # s_h = xlwt.easyxf("font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center")
        # print("---------------------------------writting into sheet--------------------")
        ws.cell(row = 1, column = 1).value = 'Portal Name'
        ws.cell(row = 1, column = 2).value = 'Product Name'
        ws.cell(row = 1, column = 3).value = 'Product Category'
        ws.cell(row = 1, column = 4).value = 'Sales'
        ws.cell(row = 1, column = 5).value = 'Total'
        ws.cell(row = 1, column = 6).value = 'Month'
        ws.cell(row = 1, column = 7).value = 'Quantity'
        invoice_ids = self.env['account.move'].search([], order='id asc')
        months = {'jan':'January', 'feb':'February', 'mar':'March', 'apr':'April', 'may':'May', 'jun':'June', 'jul':'July', 'aug':'August', 'sep':'September', 'oct':'October', 'nov':'November', 'dec':'December'}
        if invoice_ids:
            row = 2
            for inv in invoice_ids:
                for inv_l in inv.invoice_line_ids:
                    ws.cell(row, 1).value = inv.portal_id.name
                    ws.cell(row, 2).value = inv_l.product_id.name
                    ws.cell(row, 3).value = inv_l.product_id.categ_id.name
                    ws.cell(row, 4).value = inv_l.c_price_subtotal
                    ws.cell(row, 5).value = inv_l.grand_subtotal
                    ws.cell(row, 6).value = months[inv.month_of_sale]
                    ws.cell(row, 7).value = inv_l.quantity
                    row+=1    
        else:
            raise Warning("Currently No Sales Order For This Data!!")
        filename = "/tmp/All_Order.xlsx"
        workbook.save(filename)
        file = open(filename, "rb")
        file_data = file.read()
        out = base64.encodebytes(file_data)
        self.write({'file_name': filename, 'file_xls':out, 'flag': True})
        return {
           'type': 'ir.actions.act_window',
           'res_model': 'xls.report',
           'view_mode': 'form',
           'view_type': 'form',
           'res_id': self.id,
           'target': 'new',
        }
    # def action_all_order(self):
    #     inv_ids=self.env['account.move'].search([], order='id asc')
    #     workbook = openpyxl.Workbook()
    #     # wb = workbook.create_sheet('Sheet1')
    #     ws = workbook.active
    #     # s_h = xlwt.easyxf("font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center")
    #     # print("---------------------------------writting into sheet--------------------")
    #     ws.cell(row = 1, column = 1).value = 'Date'
    #     ws.cell(row = 1, column = 2).value = 'Invoice number'
    #     ws.cell(row = 1, column = 3).value = 'Channel entry'
    #     ws.cell(row = 1, column = 4).value = 'Channel Ledger'
    #     ws.cell(row = 1, column = 5).value = 'Product Name'
    #     ws.cell(row = 1, column = 6).value = 'Product SKU Code'
    #     ws.cell(row = 1, column = 7).value = 'Qty'
    #     ws.cell(row = 1, column = 8).value = 'Unit Price'
    #     ws.cell(row = 1, column = 9).value = 'Product Category'
    #     ws.cell(row = 1, column = 10).value = 'conversion rate'
    #     ws.cell(row = 1, column = 11).value = 'Total'
    #     ws.cell(row = 1, column = 12).value = 'Customer Name'
    #     ws.cell(row = 1, column = 13).value = 'Shipping Address Name'
    #     ws.cell(row = 1, column = 14).value = 'Shipping Address Line 1'
    #     ws.cell(row = 1, column = 15).value = 'Shipping Address Line 2'
    #     ws.cell(row = 1, column = 16).value = 'Shipping Address City'
    #     ws.cell(row = 1, column = 17).value = 'Shipping Address State'
    #     ws.cell(row = 1, column = 18).value = 'Shipping Address Country'
    #     ws.cell(row = 1, column = 19).value = 'Shipping Address Pincode'
    #     ws.cell(row = 1, column = 20).value = 'Shipping Address Phone'
    #     ws.cell(row = 1, column = 21).value = 'Shipping Provider'
    #     ws.cell(row = 1, column = 22).value = 'AWB num'
    #     ws.cell(row = 1, column = 23).value = 'Sales'
    #     ws.cell(row = 1, column = 24).value = 'Sales Ledger'
    #     ws.cell(row = 1, column = 25).value = 'CGST'
    #     ws.cell(row = 1, column = 26).value = 'CGST Rate'
    #     ws.cell(row = 1, column = 27).value = 'SGST'
    #     ws.cell(row = 1, column = 28).value = 'SGST Rate'
    #     ws.cell(row = 1, column = 29).value = 'IGST'
    #     ws.cell(row = 1, column = 30).value = 'IGST Rate'
    #     ws.cell(row = 1, column = 31).value = 'UTGST'
    #     ws.cell(row = 1, column = 32).value = 'UTGST Rate'
    #     ws.cell(row = 1, column = 33).value = 'CESS'
    #     ws.cell(row = 1, column = 34).value = 'CESS Rate'
    #     ws.cell(row = 1, column = 35).value = 'Other charges'
    #     ws.cell(row = 1, column = 36).value = 'Other charges Ledger'
    #     ws.cell(row = 1, column = 37).value = 'Other charges1'
    #     ws.cell(row = 1, column = 38).value = 'Other charges Ledger1'
    #     ws.cell(row = 1, column = 39).value = 'Service tax'
    #     ws.cell(row = 1, column = 40).value = 'ST Ledger'
    #     ws.cell(row = 1, column = 41).value = 'IMEI'
    #     ws.cell(row = 1, column = 42).value = 'Godown'
    #     ws.cell(row = 1, column = 43).value = 'Dispatch Date/Cancellation Date'
    #     ws.cell(row = 1, column = 44).value = 'Narration'
    #     ws.cell(row = 1, column = 45).value = 'Entity'
    #     ws.cell(row = 1, column = 46).value = 'Voucher Type Name'
    #     ws.cell(row = 1, column = 47).value = 'TIN NO'
    #     ws.cell(row = 1, column = 48).value = 'Original Invoice Date'
    #     ws.cell(row = 1, column = 49).value = 'Original Sale No'
    #     ws.cell(row = 1, column = 50).value = 'odoo order no'
    #     ws.cell(row = 1, column = 51).value = 'Suborder no'
    #     ws.cell(row = 1, column = 52).value = 'HSN'
    #     ws.cell(row = 1, column = 53).value = 'Status'
    #     ws.cell(row = 1, column = 54).value = 'Shipping From State'
    #     ws.cell(row = 1, column = 55).value = 'Original Order'
    #     ws.cell(row = 1, column = 56).value = 'Shipping Charges'
    #     ws.cell(row = 1, column = 57).value = 'Portal Name'
    #     ws.cell(row = 1, column = 58).value = 'Month'
    #     # invoice_ids = None
    #     # if self.month_of_sale:
    #     #     # print("-----------self.month_of_sale:----", self.month_of_sale)
    #     #     # print("-----------invoice_ids----", invoice_ids)
    #     #     invoice_ids = self.env['account.move'].search([('month_of_sale', '=', self.month_of_sale), ('state', '=', self.invoice_state)], order='id asc')
    #     # if self.start_date and self.end_date:
    #     #     invoice_ids = self.env['account.move'].search([('invoice_date', '>=', self.start_date), ('invoice_date', '<=', self.end_date), ('state', '=', self.invoice_state)], order='id asc')
    #     # print("-------------invoice_ids--------------", invoice_ids)
    #     invoice_ids = self.env['account.move'].search([], order='id asc')
    #     if invoice_ids:
    #         row = 2
    #         for inv in invoice_ids:
    #             for inv_l in inv.invoice_line_ids:
    #                 ws.cell(row, 1).value = inv.invoice_date.strftime("%d-%m-%Y")
    #                 ws.cell(row, 2).value = inv.invoice_number
    #                 ws.cell(row, 3).value = inv.portal_id.name
    #                 ws.cell(row, 4).value = inv.portal_id.name
    #                 ws.cell(row, 5).value = inv_l.product_id.name
    #                 ws.cell(row, 6).value = inv_l.product_id.default_code
    #                 ws.cell(row, 7).value = inv_l.quantity
    #                 ws.cell(row, 8).value = inv_l.price_unit
    #                 ws.cell(row, 9).value = inv_l.product_id.categ_id.name
    #                 ws.cell(row, 10).value = 1
    #                 ws.cell(row, 11).value = inv_l.grand_subtotal
    #                 ws.cell(row, 12).value = inv.partner_id.name
    #                 ws.cell(row, 13).value = inv.partner_id.name
    #                 ws.cell(row, 14).value = ''
    #                 ws.cell(row, 15).value = ''
    #                 ws.cell(row, 16).value = inv.partner_id.city
    #                 ws.cell(row, 17).value = inv.partner_id.state_id.l10n_in_tin+'-'+inv.partner_id.state_id.name
    #                 ws.cell(row, 18).value = inv.partner_id.country_id.name
    #                 ws.cell(row, 19).value = inv.partner_id.zip
    #                 ws.cell(row, 20).value = ''
    #                 ws.cell(row, 21).value = ''
    #                 ws.cell(row, 22).value = ''
    #                 ws.cell(row, 23).value = inv_l.c_price_subtotal
    #                 ws.cell(row, 24).value = 'CGST+SGST Sale-'+'B2C'+'-'+str(int(inv_l.tax_ids[0].amount))+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else 'IGST Sale-'+'B2C'+'-'+str(int(inv_l.tax_ids[0].amount))+'%'
    #                 ws.cell(row, 25).value = inv_l.cgst_amount
    #                 cgst = inv_l.tax_ids[0].children_tax_ids[0].amount
    #                 ws.cell(row, 26).value = 'B2C-Output Tax -CGST@ '+str(cgst)+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else ''
    #                 ws.cell(row, 27).value = inv_l.sgst_amount
    #                 ws.cell(row, 28).value = 'B2C-Output Tax -SGST@ '+str(cgst)+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else ''
    #                 ws.cell(row, 29).value = inv_l.igst_amount
    #                 ws.cell(row, 30).value = 'B2C-Output Tax -IGST@ '+str(int(inv_l.tax_ids[0].amount))+'%' if inv.partner_id.state_id.name != inv.warehouse_id.state_id.name else ''
    #                 ws.cell(row, 31).value = ''
    #                 ws.cell(row, 32).value = ''
    #                 ws.cell(row, 33).value = ''
    #                 ws.cell(row, 34).value = ''
    #                 ws.cell(row, 35).value = ''
    #                 ws.cell(row, 36).value = ''
    #                 ws.cell(row, 37).value = ''
    #                 ws.cell(row, 38).value = ''
    #                 ws.cell(row, 39).value = ''
    #                 ws.cell(row, 40).value = ''
    #                 ws.cell(row, 41).value = ''
    #                 ws.cell(row, 42).value = inv.warehouse_id.name
    #                 ws.cell(row, 43).value = ''
    #                 ws.cell(row, 44).value = 'Being order id '+str(inv.sale_number)+' Invoice no '+str(inv.invoice_number)+' suborder id '+str(inv.order_id)+' sale through '+str(inv.order_type)+' customer name '+str(inv.partner_id.name)+' state '+str(inv.partner_id.state_id.name)+' Godown '+str(inv.warehouse_id.name)
    #                 ws.cell(row, 45).value = 'NEW INVOICE'
    #                 ws.cell(row, 46).value = 'Sale-CGST+SGST-B2C' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else 'Sale-IGST-B2C'
    #                 ws.cell(row, 47).value = ''
    #                 ws.cell(row, 48).value = inv.invoice_date.strftime("%d-%m-%Y")
    #                 ws.cell(row, 49).value = inv.invoice_number
    #                 ws.cell(row, 50).value = inv.sale_number
    #                 ws.cell(row, 51).value = inv.order_id
    #                 ws.cell(row, 52).value = inv_l.product_id.l10n_in_hsn_code
    #                 ws.cell(row, 53).value = 'INVOICED' if inv.state == 'posted' else ''
    #                 ws.cell(row, 54).value = inv.warehouse_id.state_id.name
    #                 ws.cell(row, 55).value = inv.original_order_id
    #                 ws.cell(row, 56).value = inv_l.shipping_charges
    #                 ws.cell(row, 57).value = inv.portal_id.name
    #                 ws.cell(row, 58).value = inv.month_of_sale
    #                 row+=1    
    #     else:
    #         raise Warning("Currently No Sales Order For This Data!!")
    #     filename = "/tmp/All_Order.xlsx"
    #     workbook.save(filename)
    #     file = open(filename, "rb")
    #     file_data = file.read()
    #     out = base64.encodebytes(file_data)
    #     self.write({'file_name': filename, 'file_xls':out, 'flag': True})
    #     return {
    #        'type': 'ir.actions.act_window',
    #        'res_model': 'xls.report',
    #        'view_mode': 'form',
    #        'view_type': 'form',
    #        'res_id': self.id,
    #        'target': 'new',
    #     }
    def action_confirm_inv(self):
        action_confirm_inv_draft=self.env['account.move'].search([('state', '=', 'draft')], order='id asc')
        print("------------------action_confirm_inv_draft----", action_confirm_inv_draft)
        for ioaf in action_confirm_inv_draft:
            ioaf.action_post()
            # break;
    def action_xls(self):
        # print("============action_report================")
        workbook = xlwt.Workbook()
        ws = workbook.add_sheet('Sheet1')
        s_h = xlwt.easyxf("font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center")
        print("---------------------------------writting into sheet--------------------")
        ws.write(0, 0, 'Date', s_h)
        ws.write(0, 1, 'Sale Order Number', s_h)
        ws.write(0, 2, 'Invoice number', s_h)
        ws.write(0, 3, 'Channel entry', s_h)
        ws.write(0, 4, 'Channel Ledger', s_h)
        ws.write(0, 5, 'Product Name', s_h)
        ws.write(0, 6, 'Product SKU Code', s_h)
        ws.write(0, 7, 'Qty', s_h)
        ws.write(0, 8, 'Unit Price', s_h)
        ws.write(0, 9, 'Currency', s_h)
        ws.write(0, 10, 'conversion rate', s_h)
        ws.write(0, 11, 'Total', s_h)
        ws.write(0, 12, 'Customer Name', s_h)
        ws.write(0, 13, 'Shipping Address Name', s_h)
        ws.write(0, 14, 'Shipping Address Line 1', s_h)
        ws.write(0, 15, 'Shipping Address Line 2', s_h)
        ws.write(0, 16, 'Shipping Address City', s_h)
        ws.write(0, 17, 'Shipping Address State', s_h)
        ws.write(0, 18, 'Shipping Address Country', s_h)
        ws.write(0, 19, 'Shipping Address Pincode', s_h)
        ws.write(0, 20, 'Shipping Address Phone', s_h)
        ws.write(0, 21, 'Shipping Provider', s_h)
        ws.write(0, 22, 'AWB num', s_h)
        ws.write(0, 23, 'Sales', s_h)
        ws.write(0, 24, 'Sales Ledger', s_h)
        ws.write(0, 25, 'CGST', s_h)
        ws.write(0, 26, 'CGST Rate', s_h)
        ws.write(0, 27, 'SGST', s_h)
        ws.write(0, 28, 'SGST Rate', s_h)
        ws.write(0, 29, 'IGST', s_h)
        ws.write(0, 30, 'IGST Rate', s_h)
        ws.write(0, 31, 'UTGST', s_h)
        ws.write(0, 32, 'UTGST Rate', s_h)
        ws.write(0, 33, 'CESS', s_h)
        ws.write(0, 34, 'CESS Rate', s_h)
        ws.write(0, 35, 'Other charges', s_h)
        ws.write(0, 36, 'Other charges Ledger', s_h)
        ws.write(0, 37, 'Other charges1', s_h)
        ws.write(0, 38, 'Other charges Ledger1', s_h)
        ws.write(0, 39, 'Service tax', s_h)
        ws.write(0, 40, 'ST Ledger', s_h)
        ws.write(0, 41, 'IMEI', s_h)
        ws.write(0, 42, 'Godown', s_h)
        ws.write(0, 43, 'Dispatch Date/Cancellation Date', s_h)
        ws.write(0, 44, 'Narration', s_h)
        ws.write(0, 45, 'Entity', s_h)
        ws.write(0, 46, 'Voucher Type Name', s_h)
        ws.write(0, 47, 'TIN NO', s_h)
        ws.write(0, 48, 'Original Invoice Date', s_h)
        ws.write(0, 49, 'Original Sale No', s_h)
        ws.write(0, 50, 'odoo order no', s_h)
        ws.write(0, 51, 'Suborder no', s_h)
        ws.write(0, 52, 'HSN', s_h)
        ws.write(0, 53, 'Status', s_h)
        ws.write(0, 54, 'Shipping From State', s_h)
        ws.write(0, 55, 'Original Order', s_h)
        ws.write(0, 56, 'Shipping Charges', s_h)
        ws.write(0, 57, 'Portal Name', s_h)
        ws.write(0, 58, 'Order Date(Portal)', s_h)
        ws.write(0, 59, 'Invoice Date(Portal)', s_h)
        # ws.write(0, 57, 'Gift Wrap Price', s_h)
        # ws.write(0, 58, 'Promotional Discount', s_h)
        invoice_ids = None
        if self.month_of_sale:
            # print("-----------self.month_of_sale:----", self.month_of_sale)
            invoice_ids = self.env['account.move'].search([('month_of_sale', '=', self.month_of_sale), ('state', '=', self.invoice_state)], order='id asc')
            # print("-----------invoice_ids----", invoice_ids)
        if self.start_date and self.end_date:
            invoice_ids = self.env['account.move'].search([('invoice_date', '>=', self.start_date), ('invoice_date', '<=', self.end_date), ('state', '=', self.invoice_state)], order='id asc')
        # print("-------------invoice_ids--------------", invoice_ids)
        if invoice_ids:
            row = 1
            for inv in invoice_ids:
                if inv.so_id.order_date:
                    o_date = inv.so_id.order_date.strftime("%d-%m-%Y")
                else:
                    o_date = ''
                for inv_l in inv.invoice_line_ids:
                    ws.write(row, 0, inv.invoice_date.strftime("%d-%m-%Y"))
                    ws.write(row, 1, inv.invoice_number)
                    ws.write(row, 2, inv.invoice_number)
                    ws.write(row, 3, inv.portal_id.name)
                    ws.write(row, 4, inv.portal_id.name)
                    ws.write(row, 5, inv_l.product_id.name)
                    ws.write(row, 6, inv_l.product_id.default_code)
                    ws.write(row, 7, inv_l.quantity)
                    ws.write(row, 8, inv_l.price_unit)
                    ws.write(row, 9, 'INR')
                    ws.write(row, 10, 1)
                    ws.write(row, 11, inv_l.grand_subtotal)
                    ws.write(row, 12, inv.partner_id.name)
                    ws.write(row, 13, inv.partner_id.name)
                    ws.write(row, 14, '')
                    ws.write(row, 15, '')
                    # ws.write(row, 14, inv.partner_id.street)
                    # ws.write(row, 15, inv.partner_id.street2 if inv.partner_id.street2 else '')
                    ws.write(row, 16, inv.partner_id.city)
                    ws.write(row, 17, inv.partner_id.state_id.l10n_in_tin+'-'+inv.partner_id.state_id.name)
                    ws.write(row, 18, inv.partner_id.country_id.name)
                    ws.write(row, 19, inv.partner_id.zip)
                    ws.write(row, 20, '')
                    ws.write(row, 21, '')
                    ws.write(row, 22, '')
                    ws.write(row, 23, inv_l.c_price_subtotal)
                    ws.write(row, 24, 'CGST+SGST Sale-'+'B2C'+'-'+str(int(inv_l.tax_ids[0].amount))+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else 'IGST Sale-'+'B2C'+'-'+str(int(inv_l.tax_ids[0].amount))+'%')
                    # ws.write(row, 24, 'CGST+SGST Sale-'+str(inv.order_category).upper()+'-'+str(int(inv_l.tax_ids[0].amount))+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else 'IGST Sale-'+str(inv.order_category).upper()+'-'+str(int(inv_l.tax_ids[0].amount))+'%')
                    ws.write(row, 25, inv_l.cgst_amount)
                    cgst = inv_l.tax_ids[0].children_tax_ids[0].amount
                    # cgst = inv_l.tax_ids[0].amount/2 if (inv_l.tax_ids[0].amount/2)%1>0.0 else int(inv_l.tax_ids[0].amount/2)
                    # cgst = inv_l.tax_ids[0].amount/2 if (inv_l.tax_ids[0].amount/2)%1>0.0 else int(inv_l.tax_ids[0].amount/2)
                    ws.write(row, 26, 'B2C-Output Tax -CGST@ '+str(cgst)+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else '')
                    # ws.write(row, 26, str(inv.order_category).upper()+'-Output Tax -CGST@ '+str(cgst)+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else '')
                    ws.write(row, 27, inv_l.sgst_amount)
                    ws.write(row, 28, 'B2C-Output Tax -SGST@ '+str(cgst)+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else '')
                    # ws.write(row, 28, str(inv.order_category).upper()+'-Output Tax -SGST@ '+str(cgst)+'%' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else '')
                    ws.write(row, 29, inv_l.igst_amount)
                    ws.write(row, 30, 'B2C-Output Tax -IGST@ '+str(int(inv_l.tax_ids[0].amount))+'%' if inv.partner_id.state_id.name != inv.warehouse_id.state_id.name else '')
                    # ws.write(row, 30, str(inv.order_category).upper()+'-Output Tax -IGST@ '+str(int(inv_l.tax_ids[0].amount))+'%' if inv.partner_id.state_id.name != inv.warehouse_id.state_id.name else '')
                    ws.write(row, 31, '')
                    ws.write(row, 32, '')
                    ws.write(row, 33, '')
                    ws.write(row, 34, '')
                    # ws.write(row, 35, inv_l.other_charges)
                    ws.write(row, 35, '')
                    ws.write(row, 36, '')
                    # ws.write(row, 36, 'Shipping Charges on Sale CGST+SGST' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else 'Shipping Charges on Sale IGST')
                    ws.write(row, 37, '')
                    ws.write(row, 38, '')
                    ws.write(row, 39, '')
                    ws.write(row, 40, '')
                    ws.write(row, 41, '')
                    ws.write(row, 42, inv.warehouse_id.name)
                    ws.write(row, 43, '')
                    ws.write(row, 44, 'Being order id '+str(inv.sale_number)+' Invoice no '+str(inv.invoice_number)+' suborder id '+str(inv.order_id)+' sale through '+str(inv.order_type)+' customer name '+str(inv.partner_id.name)+' state '+str(inv.partner_id.state_id.name)+' Godown '+str(inv.warehouse_id.name))
                    ws.write(row, 45, 'NEW INVOICE')
                    ws.write(row, 46, 'Sale-CGST+SGST-B2C' if inv.partner_id.state_id.name == inv.warehouse_id.state_id.name else 'Sale-IGST-B2C')
                    ws.write(row, 47, '')
                    ws.write(row, 48, inv.invoice_date.strftime("%d-%m-%Y"))
                    ws.write(row, 49, inv.invoice_number)
                    ws.write(row, 50, inv.sale_number)
                    ws.write(row, 51, inv.order_id)
                    ws.write(row, 52, inv_l.product_id.l10n_in_hsn_code)
                    ws.write(row, 53, 'INVOICED' if inv.state == 'posted' else '')
                    ws.write(row, 54, inv.warehouse_id.state_id.name)
                    ws.write(row, 55, inv.original_order_id)
                    ws.write(row, 56, inv_l.shipping_charges)
                    ws.write(row, 57, inv.portal_id.name)
                    ws.write(row, 58, o_date)
                    ws.write(row, 59, inv.so_id.invoice_date.strftime("%d-%m-%Y") if inv.so_id.invoice_date else '')
                    row+=1    
        else:
            raise Warning("Currently No Sales Order For This Data!!")
        filename = '/tmp/Reports'+ '.xls'
        workbook.save(filename)
        file = open(filename, "rb")
        file_data = file.read()
        out = base64.encodebytes(file_data)
        self.write({'file_name': filename, 'file_xls':out, 'flag': True})
        return {
           'type': 'ir.actions.act_window',
           'res_model': 'xls.report',
           'view_mode': 'form',
           'view_type': 'form',
           'res_id': self.id,
           'target': 'new',
        }

# class CredtNotesXls(models.TransientModel):
#     _name = "credit.notes.xls"
#     _description = 'Credt Notes Xls'
#     start_date = fields.Date(string='Date From', required=True)
#     end_date = fields.Date(string='Date To', required=True)
#     file_name = fields.Char('Name', size=256)
#     file_xls = fields.Binary(' Report', readonly=True)
#     flag = fields.Boolean(string="Flag")
#     def action_credit_note_xls(self):
#         print("============action_credit_note_xls================")
#         workbook = xlwt.Workbook()
#         ws = workbook.add_sheet('Sheet1')
#         s_h = xlwt.easyxf("font: bold on; pattern: pattern solid, fore_colour yellow; align: horiz center")
#         print("---------------------------------writting into sheet--------------------")
#         ws.write(0, 0, 'Date', s_h)
#         ws.write(0, 1, 'Sale Order Number', s_h)
#         ws.write(0, 2, 'Invoice number', s_h)
#         ws.write(0, 3, 'Channel entry', s_h)
#         ws.write(0, 4, 'Channel Ledger', s_h)
#         ws.write(0, 5, 'Product Name', s_h)
#         ws.write(0, 6, 'Product SKU Code', s_h)
#         ws.write(0, 7, 'Qty', s_h)
#         ws.write(0, 8, 'Unit Price', s_h)
#         ws.write(0, 9, 'Currency', s_h)
#         ws.write(0, 10, 'conversion rate', s_h)
#         ws.write(0, 11, 'Total', s_h)
#         ws.write(0, 12, 'Sales', s_h)
#         ws.write(0, 13, 'Sales Ledger', s_h)
#         ws.write(0, 14, 'CGST', s_h)
#         ws.write(0, 15, 'CGST Rate', s_h)
#         ws.write(0, 16, 'SGST', s_h)
#         ws.write(0, 17, 'SGST Rate', s_h)
#         ws.write(0, 18, 'IGST', s_h)
#         ws.write(0, 19, 'IGST Rate', s_h)
#         ws.write(0, 20, 'UTGST', s_h)
#         ws.write(0, 21, 'UTGST Rate', s_h)
#         ws.write(0, 22, 'CESS', s_h)
#         ws.write(0, 23, 'CESS Rate', s_h)
#         ws.write(0, 24, 'Other charges', s_h)
#         ws.write(0, 25, 'Other charges Ledger', s_h)
#         ws.write(0, 26, 'Other charges1', s_h)
#         ws.write(0, 27, 'Other charges Ledger1', s_h)
#         ws.write(0, 28, 'Service tax', s_h)
#         ws.write(0, 29, 'ST Ledger', s_h)
#         ws.write(0, 30, 'IMEI', s_h)
#         ws.write(0, 31, 'Godown', s_h)
#         ws.write(0, 32, 'Dispatch Date/Cancellation Date', s_h)
#         ws.write(0, 33, 'Narration', s_h)
#         ws.write(0, 34, 'Entity', s_h)
#         ws.write(0, 35, 'Voucher Type Name', s_h)
#         ws.write(0, 36, 'TIN NO', s_h)
#         ws.write(0, 37, 'Original Invoice Date', s_h)
#         ws.write(0, 38, 'Original Sale No', s_h)
#         ws.write(0, 39, 'odoo order no', s_h)
#         ws.write(0, 40, 'Suborder no', s_h)
#         ws.write(0, 41, 'HSN', s_h)
#         ws.write(0, 42, 'Status', s_h)
#         ws.write(0, 43, 'Mode of payment', s_h)
#         ws.write(0, 44, 'GSTIN No.', s_h)
#         ws.write(0, 45, 'Cost Price', s_h)
#         ws.write(0, 46, 'Total Cost', s_h)
#         # s_o_ids = self.env['sale.order'].search([('date_order', '>=', self.start_date), ('date_order', '<=', self.end_date), ('state', '=', self.order_state)], order='id asc')
#         # if s_o_ids:
#         #     row = 1
#         #     for so in s_o_ids:
#         #         ws.write(row, 0, so.name)
#         #     row+=1    
#         # else:
#         #     raise Warning("Currently No Sales Order For This Data!!")


#         filename = 'CREDIT NOTES'+'.xls'
#         workbook.save(filename)
#         file = open(filename, "rb")
#         file_data = file.read()
#         out = base64.encodestring(file_data)
#         self.write({'file_name': filename, 'file_xls':out, 'flag': True})
#         return {
#            'type': 'ir.actions.act_window',
#            'res_model': 'credit.notes.xls',
#            'view_mode': 'form',
#            'view_type': 'form',
#            'res_id': self.id,
#            'target': 'new',
#         }
